from aiogram import Router, F, Bot
from aiogram.filters import Command
from aiogram.types import Message, CallbackQuery, FSInputFile
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import StatesGroup, State

from config import BOT_TOKEN, ADMIN_CHAT_ID

import openpyxl
import os
import datetime 

from keyboards.keyboards import start_keyboard, personal_account_keyboard, admin_keyboard
from keyboards.keyboards_tools import create_list_products, create_list_busket

from WorkWithDataBase import WorkWithDatabase

db = WorkWithDatabase()

bot = Bot(token = BOT_TOKEN)
router = Router()

class Flag(StatesGroup):
    wait_for_file = State()
    wait_for_count = State()
    wait_for_edit_count = State()

def get_full_user_name(user):
    full_name = user.first_name if user.first_name else "Неизвестный"
    if user.last_name:
        full_name += f" {user.last_name}"
    return full_name

@router.message(Command('start'))
async def start_command_handler(message: Message):
    text = 'Добро пожаловать в бот! 🎉\n\nЗдесь вы можете заказать принтеры различного вида, с помощью удобной системы. \n\n - Для выбора товаров и добавления их в корзину перейдите по кнопке "Общая выпскаемая продукция"\n\n - Для перехода в личный кабинет перейдите по кнопке: "Личный кабинет"'
    await message.answer(text, reply_markup = start_keyboard)

@router.message(Command('admin_panel'))
async def admin_panel_command_handler(message: Message):
    text = 'Добро пожаловать в admin panel'
    await message.answer(text, reply_markup = admin_keyboard)

@router.callback_query(F.data == 'cabinet_basket')
async def cabinet_basket_handler(c: CallbackQuery):
    await c.message.edit_text("Добро пожаловать в личный кабинет! 👋 \n - для просмотра содержимого корзины перейдите по кнопке 'Корзина'\n - для оплаты и заказа выбранных товаров перейдите по кнопке 'Оплатить товары'\n - для очистки содержимого корзины перейдите по кнопке 'Очистить корзину'\n - Для возвращения к выбору товаров перейдите по кнопке 'На главную страницу' ", reply_markup = personal_account_keyboard)

@router.callback_query(F.data == 'production')
async def show_production_handler(c: CallbackQuery):
    products_data = db.show_products()
    products_keyboard = await create_list_products(products_data)
    await c.message.edit_text(
        "Выберите товар который вы хотели бы заказать:", 
        reply_markup = products_keyboard
    )

@router.callback_query(F.data.startswith('add_to_busket'))
async def add_to_busket(c: CallbackQuery, state: FSMContext):
    try:
        id_product = int(c.data.split(':')[1])
        product_name = db.show_product_for_id(id_product)[0]
        
        await state.update_data(id_product = id_product)
        await c.answer(f"Вы выбрали: {product_name}", show_alert=False)
        await c.message.edit_text(f'Вы выбрали {product_name}. Введите желаемое количество товара:')
        await state.set_state(Flag.wait_for_count)
    except Exception as e:
        await c.answer("Произошла ошибка при выборе товара. Попробуйте снова.", show_alert=True)

@router.message(Flag.wait_for_count)
async def get_count_handler(message: Message, state: FSMContext):
    try:
        count = int(message.text)
        if count <= 0:
            raise ValueError
    except ValueError:
        await message.answer("Неверный формат. Введите целое положительное число.")
        return

    data = await state.get_data()
    id_product = data.get('id_product')
    chat_id = str(message.chat.id)

    db.insert_into_user_basket(chat_id, id_product, count)
    
    await state.clear()
    await message.answer("Товар  успешно добавлен в корзину!🎉\n\nЕсли хотите добавить еще товары в корзину перейдите по кнопке: 'Общая выпускаемая продукция'\n\nДля просмотра корзины и оплаты выьранных товаров перейдите по кнопке 'Личный кабинет' ", reply_markup = start_keyboard)

@router.callback_query(F.data == 'show_basket')
async def show_basket_handler(c: CallbackQuery):
    chat_id = str(c.message.chat.id)
    basket = db.select_busket(chat_id)

    if not basket:
        await c.message.edit_text("Ваша корзина пуста. ☺️", reply_markup = personal_account_keyboard)
        return

    basket_keyboard = await create_list_busket(basket)
    await c.message.edit_text("Ваша корзина:", reply_markup = basket_keyboard)

@router.callback_query(F.data == 'order_busket')
async def order_busket(c: CallbackQuery):
    chat_id = str(c.message.chat.id)
    customer_name = get_full_user_name(c.from_user)
    
    basket_items = db.select_busket(chat_id)
    
    if not basket_items:
        await c.answer("Ваша корзина пуста. ☺️", show_alert=True)
        return

    now = datetime.datetime.now()
    order_date = now.strftime("%Y-%m-%d %H:%M:%S")

    wb = openpyxl.Workbook()
    sheet = wb.active

    sheet['A1'] = "Дата заказа:"
    sheet['B1'] = order_date
    sheet['A2'] = "Заказчик:"
    sheet['B2'] = customer_name
    
    start_row = 5
    headers = ["Имя товара", "Количество товара (шт.)", "Цена за штуку, рубли", "Суммарная стоимость, рубли"]
    
    for col_num, header in enumerate(headers, start=1):
        sheet.cell(row=start_row, column=col_num, value=header)

    current_row = start_row + 1
    total_order_cost = 0

    try:
        for item in basket_items:
            product_id = item[2]
            count = item[3]
            
            product_info = db.show_product_for_id(product_id)
            product_name = product_info[0]
            product_cost = product_info[1]
            
            item_total_cost = product_cost * count
            total_order_cost += item_total_cost
            
            row_data = [
                product_name, 
                count,
                product_cost,
                item_total_cost
            ]
            
            for col_num, value in enumerate(row_data, start=1):
                sheet.cell(row=current_row, column=col_num, value=value)
            
            current_row += 1

    except Exception as e:
        await c.answer(f"Ошибка при обработке одного из товаров корзины: {e}. Заказ отменен.", show_alert=True)
        return
        
    total_row = current_row + 2 
    sheet.cell(row=total_row, column=1, value="ОБЩАЯ СУММА ЗАКАЗА:")
    sheet.cell(row=total_row, column=2, value=total_order_cost)
    sheet.cell(row=total_row, column=3, value="рубли")
    
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                cell_value_str = str(cell.value) if cell.value is not None else ""
                if len(cell_value_str) > max_length:
                    max_length = len(cell_value_str)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width


    excel_file_name = f"Заказ_{chat_id}_{now.strftime('%Y%m%d%H%M%S')}.xlsx"
    wb.save(excel_file_name)
    
    input_file = FSInputFile(excel_file_name)
    caption_text = f"Новый заказ от {customer_name} ({chat_id}).\nДата: {order_date}"
    
    try:
        await bot.send_document(chat_id=ADMIN_CHAT_ID, document=input_file, caption=caption_text)
        await c.answer("Заказ отправлен! 🎉", show_alert=False)
    except Exception as e:
        await c.answer("Ошибка отправки администратору, но ваш заказ оформлен.", show_alert=True)
    
    await c.message.answer_document(document = input_file, caption = f"Ваш заказ от {order_date} успешно оформлен.")
    
    await c.message.answer("Чек во вложении")
    
    db.delete_busket(chat_id)
    os.remove(excel_file_name)

@router.callback_query(F.data == 'clear_all_basket')
async def clear_all_basket_handler(c: CallbackQuery):
    chat_id = str(c.message.chat.id)
    db.delete_busket(chat_id)
    await c.answer("Корзина полностью очищена! ✅", show_alert=True)
    await c.message.edit_text("Добро пожаловать в личный кабинет! 👋 \n - для просмотра содержимого корзины перейдите по кнопке 'Корзина'\n - для оплаты и заказа выбанных товаров перейдите по кнопке 'Оплатить товары'\n - для очистки содержимого корзины перейдите по кнопке 'Очистить корзину'\n - Для возвращения к выбору товаров перейдите по кнопке 'На главную страницу' ", reply_markup = personal_account_keyboard)

@router.callback_query(F.data.startswith('delete_item'))
async def delete_item_handler(c: CallbackQuery):
    item_id = int(c.data.split(':')[1])
    db.delete_item_busket(item_id)
    
    await c.answer("Элемент удален из корзины.", show_alert=False)
    await show_basket_handler(c)

@router.callback_query(F.data.startswith('edit_item'))
async def edit_item_handler(c: CallbackQuery, state: FSMContext):
    item_id = int(c.data.split(':')[1])
    
    await state.update_data(item_id_to_edit=item_id)
    
    await c.answer("Введите новое количество.", show_alert=False)
    await c.message.edit_text("Введите новое количество для выбранного товара:")
    await state.set_state(Flag.wait_for_edit_count)

@router.message(Flag.wait_for_edit_count)
async def get_edit_count_handler(message: Message, state: FSMContext):
    try:
        new_count = int(message.text)
        if new_count <= 0:
            raise ValueError
    except ValueError:
        await message.answer("Неверный формат. Введите целое положительное число.")
        return

    data = await state.get_data()
    item_id_to_edit = data.get('item_id_to_edit')
    
    db.update_item_count(item_id_to_edit, new_count)
    
    await state.clear()
    await message.answer("Количество товара обновлено!", reply_markup = personal_account_keyboard)

@router.callback_query(F.data == 'back_to_start')
async def back_to_start_handler(c: CallbackQuery):
    await c.message.edit_text('Добро пожаловать в бот! 👋\n\n Здесь вы можете заказать принтеры различного вида, с помощью удобной системы. \n\n - Для выбора товаров и добавления их в корзину перейдите по кнопке "Общая выпскаемая продукция"\n\n - Для перехода в личный кабинет перейдите по кнопке: "Личный кабинет"', reply_markup = start_keyboard)

@router.callback_query(F.data == 'file_download')
async def file_download_handler(c: CallbackQuery, state: FSMContext):
    await c.message.edit_text('Загрузите файл Excel', reply_markup = start_keyboard)
    await state.set_state(Flag.wait_for_file)

@router.message(Flag.wait_for_file)
async def process_file(message: Message, state: FSMContext):
    if not message.document:
        await message.answer("Пожалуйста, загрузите файл.")
        return
        
    file_name = message.document.file_name
    
    if not file_name.endswith('.xlsx'):
        await message.answer("Поддерживаются только файлы .xlsx.")
        return

    file_id = message.document.file_id
    file_info = await bot.get_file(file_id)
    file_path = file_info.file_path
    
    downloaded_file = await bot.download_file(file_path)
    
    with open(file_name, 'wb') as f:
        f.write(downloaded_file.read())
        
    try:
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
        
        db.delete_from_products()
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                db.save_products(str(row[0]), int(row[1]))
                
    except Exception as e:
        await message.answer(f"Ошибка при обработке файла: {e}")
    
    if os.path.exists(file_name):
        os.remove(file_name)
            
    await state.clear()
    await message.answer("Загрузка прошла успешно", reply_markup = admin_keyboard)